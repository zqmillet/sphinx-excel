"""
this module provides the directive excel.
"""

from typing import List
from docutils.parsers.rst import directives
from docutils.parsers.rst.directives.tables import RSTTable
from docutils.statemachine import StringList
from docutils.nodes import table as TableNode
from sphinx.util.docutils import SphinxDirective
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .table import Table
from .span import Span
from .coordinate import Coordinate

def get_data(worksheet: Worksheet) -> List[List[str]]:
    return [[str(cell.value) if cell.value else '' for cell in row] for row in worksheet.rows]

def get_spans(worksheet: Worksheet) -> List[Span]:
    return [
        Span(
            Coordinate(row=cell_range.min_row - 1, column=cell_range.min_col - 1),
            rows=cell_range.max_row - cell_range.min_row + 1,
            columns=cell_range.max_col - cell_range.min_col + 1,
        ) for cell_range in worksheet.merged_cell_ranges  # type: ignore
    ]

class ExcelDirective(RSTTable, SphinxDirective):
    """
    an environment for excel
    """

    option_spec = {
        **RSTTable.option_spec,
        'caption': directives.unchanged,
        'sheet': directives.unchanged,
        'headers': directives.nonnegative_int,
        'no-caption': directives.flag,
    }

    def run(self) -> List[TableNode]:
        """
        render this environment
        """

        file_path, *_ = self.arguments  # type: ignore
        _, file_path = self.env.relfn2path(file_path)
        workbook = load_workbook(file_path)
        default_worksheet_name, *_ = workbook.sheetnames
        headers = self.options.get('headers', 1)
        worksheet_name = self.options.get('sheet', default_worksheet_name)
        caption = self.options.get('caption', worksheet_name)
        no_caption = 'no-caption' in self.options

        worksheet = workbook[worksheet_name]
        table = Table(data=get_data(worksheet), spans=get_spans(worksheet), headers=headers)
        self.arguments = [] if no_caption else [caption]
        self.content = StringList(table.render().splitlines())
        return super().run()
